"""
Document Processor Module

Handles extraction of text from text-based files and conversion of images to base64.
Uses hybrid approach: text files -> extract text, images -> base64 for Vision GPT.
"""

import base64
import io
import logging
import re
from typing import Optional, Tuple, Union
from PIL import Image

logger = logging.getLogger(__name__)

# Optional imports for encoding detection
try:
    import chardet
    HAS_CHARDET = True
except ImportError:
    HAS_CHARDET = False
    logger.warning("chardet not installed. CSV encoding detection will use UTF-8 fallback.")


def _decode_content(content: Union[str, bytes], file_name: str) -> bytes:
    """Decode content from base64 string or return bytes directly."""
    if isinstance(content, str):
        if content.startswith('data:'):
            base64_part = content.split(',', 1)[1] if ',' in content else content
            return base64.b64decode(base64_part, validate=True)
        else:
            return base64.b64decode(content, validate=True)
    return content


def process_text_file(content: Union[str, bytes], file_type: str, file_name: str) -> str:
    """
    Process text-based files and extract text content.
    
    Args:
        content: File content as bytes (from FormData) or base64 string (backward compatibility)
        file_type: MIME type of the file
        file_name: Name of the file
    
    Returns:
        Extracted text content
    """
    try:
        decoded_content = _decode_content(content, file_name)
        
        # PDF
        if file_type == "application/pdf" or file_name.lower().endswith('.pdf'):
            return extract_pdf_text(decoded_content)
        
        # DOCX
        elif (
            file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" or
            file_name.lower().endswith('.docx')
        ):
            return extract_docx_text(decoded_content)
        
        # Legacy DOC
        elif file_type == "application/msword" or file_name.lower().endswith('.doc'):
            return extract_doc_text(decoded_content)
        
        # Excel XLSX
        elif (
            file_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or
            file_name.lower().endswith('.xlsx')
        ):
            return extract_xlsx_text(decoded_content)
        
        # Excel XLS (Legacy)
        elif file_type == "application/vnd.ms-excel" or file_name.lower().endswith('.xls'):
            return extract_xls_text(decoded_content)
        
        # OpenDocument Spreadsheet (ODS)
        elif file_type == "application/vnd.oasis.opendocument.spreadsheet" or file_name.lower().endswith('.ods'):
            return extract_ods_text(decoded_content)
        
        # HTML
        elif file_type == "text/html" or file_name.lower().endswith(('.html', '.htm')):
            return extract_html_text(decoded_content)
        
        # CSV
        elif file_type == "text/csv" or file_name.lower().endswith('.csv'):
            return extract_csv_text(decoded_content, file_name)
        
        # TSV
        elif file_name.lower().endswith('.tsv'):
            return extract_tsv_text(decoded_content, file_name)
        
        # Plain text
        elif file_type == "text/plain" or file_name.lower().endswith('.txt'):
            return decoded_content.decode('utf-8', errors='replace')
        
        # Fallback: try to decode as text
        else:
            return decoded_content.decode('utf-8', errors='ignore')
    
    except Exception as e:
        logger.error(f"Error processing text file {file_name}: {e}", exc_info=True)
        raise ValueError(f"Failed to extract text from {file_name}: {str(e)}")


def extract_pdf_text(pdf_content: bytes) -> str:
    """Extract text from PDF file using multiple fallback methods."""
    # Method 1: Try PyPDF2 first
    try:
        import PyPDF2
        pdf_file = io.BytesIO(pdf_content)
        pdf_reader = PyPDF2.PdfReader(pdf_file, strict=False)
        text_parts = []
        
        for page in pdf_reader.pages:
            try:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
            except Exception:
                continue
        
        if text_parts:
            return "\n\n".join(text_parts)
    except ImportError:
        pass
    except Exception:
        pass
    
    # Method 2: Fallback to PyMuPDF
    try:
        import fitz
        try:
            pdf_document = fitz.open(stream=pdf_content, filetype="pdf")
        except Exception:
            pdf_document = fitz.open(stream=io.BytesIO(pdf_content), filetype="pdf")
        
        text_parts = []
        for page_num in range(len(pdf_document)):
            try:
                page = pdf_document[page_num]
                page_text = page.get_text()
                if page_text.strip():
                    text_parts.append(page_text)
            except Exception:
                continue
        
        pdf_document.close()
        
        if text_parts:
            return "\n\n".join(text_parts)
    except ImportError:
        pass
    except Exception:
        pass
    
    raise ValueError(
        "Failed to extract text from PDF. The file may be corrupted, encrypted, "
        "or image-based (scanned PDF with no text layer)."
    )


def extract_docx_text(docx_content: bytes) -> str:
    """Extract text from DOCX file."""
    try:
        from docx import Document
        docx_file = io.BytesIO(docx_content)
        doc = Document(docx_file)
        
        text_parts = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_parts.append(paragraph.text)
        
        return "\n\n".join(text_parts)
    except ImportError:
        raise ValueError("python-docx is not installed. Please install it to process DOCX files.")
    except Exception as e:
        raise ValueError(f"Failed to extract text from DOCX: {str(e)}")


def extract_doc_text(doc_content: bytes) -> str:
    """Extract text from legacy DOC file."""
    try:
        import textract
        return textract.process(io.BytesIO(doc_content)).decode('utf-8')
    except ImportError:
        try:
            from docx2txt import process
            return process(io.BytesIO(doc_content))
        except ImportError:
            # Fallback: try to extract readable text from binary
            # This is a basic approach and may not work for all DOC files
            try:
                content_str = doc_content.decode('utf-8', errors='ignore')
                # Extract text between common markers
                text_parts = re.findall(r'[a-zA-Z0-9\s]{20,}', content_str)
                if text_parts:
                    return '\n'.join(text_parts)
                raise ValueError("Could not extract readable text from DOC file")
            except Exception:
                raise ValueError(
                    "Legacy DOC files require textract or docx2txt for proper extraction. "
                    "Install with: pip install textract or pip install docx2txt"
                )
    except Exception as e:
        raise ValueError(f"Failed to extract text from DOC: {str(e)}")


def extract_xlsx_text(xlsx_content: bytes) -> str:
    """Extract text from Excel XLSX file."""
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(xlsx_content), data_only=True)
        
        text_parts = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"\n[Sheet: {sheet_name}]\n")
            
            for row in sheet.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                if any(row_values):  # Skip empty rows
                    text_parts.append("\t".join(row_values))
        
        return "\n".join(text_parts)
    except ImportError:
        raise ValueError("openpyxl is not installed. Please install it to process XLSX files.")
    except Exception as e:
        raise ValueError(f"Failed to extract text from XLSX: {str(e)}")


def extract_xls_text(xls_content: bytes) -> str:
    """Extract text from legacy Excel XLS file."""
    try:
        import xlrd
        workbook = xlrd.open_workbook(file_contents=xls_content)
        
        text_parts = []
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            text_parts.append(f"\n[Sheet: {sheet_name}]\n")
            
            for row_idx in range(sheet.nrows):
                row_values = [str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
                if any(row_values):
                    text_parts.append("\t".join(row_values))
        
        return "\n".join(text_parts)
    except ImportError:
        raise ValueError("xlrd is not installed. Please install it to process XLS files.")
    except Exception as e:
        raise ValueError(f"Failed to extract text from XLS: {str(e)}")


def extract_ods_text(ods_content: bytes) -> str:
    """Extract text from OpenDocument Spreadsheet (ODS) file."""
    try:
        from odf.opendocument import load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        
        doc = load(io.BytesIO(ods_content))
        
        text_parts = []
        for table in doc.getElementsByType(Table):
            table_name = table.getAttribute("name") or "Sheet"
            text_parts.append(f"\n[Sheet: {table_name}]\n")
            
            for row in table.getElementsByType(TableRow):
                row_values = []
                for cell in row.getElementsByType(TableCell):
                    cell_text = ""
                    for p in cell.getElementsByType(P):
                        cell_text += p.firstChild.data if p.firstChild else ""
                    row_values.append(cell_text)
                
                if any(row_values):
                    text_parts.append("\t".join(row_values))
        
        return "\n".join(text_parts)
    except ImportError:
        raise ValueError("odfpy is not installed. Please install it to process ODS files.")
    except Exception as e:
        raise ValueError(f"Failed to extract text from ODS: {str(e)}")


def extract_html_text(html_content: bytes) -> str:
    """Extract text from HTML file."""
    try:
        from bs4 import BeautifulSoup
        
        # Try to detect encoding
        if HAS_CHARDET:
            detected = chardet.detect(html_content)
            encoding = detected.get('encoding', 'utf-8')
        else:
            encoding = 'utf-8'
        
        html_text = html_content.decode(encoding, errors='replace')
        soup = BeautifulSoup(html_text, 'lxml')
        
        # Remove script and style elements
        for script in soup(["script", "style"]):
            script.decompose()
        
        # Get text
        text = soup.get_text()
        
        # Clean up whitespace
        lines = (line.strip() for line in text.splitlines())
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        text = '\n'.join(chunk for chunk in chunks if chunk)
        
        return text
    except ImportError:
        # Fallback: basic HTML tag removal
        html_text = html_content.decode('utf-8', errors='replace')
        # Remove HTML tags using regex (simple approach)
        text = re.sub(r'<[^>]+>', '', html_text)
        return text
    except Exception as e:
        raise ValueError(f"Failed to extract text from HTML: {str(e)}")


def extract_csv_text(csv_content: bytes, file_name: str) -> str:
    """Extract text from CSV file with encoding detection."""
    if HAS_CHARDET:
        try:
            detected = chardet.detect(csv_content)
            encoding = detected.get('encoding', 'utf-8')
            confidence = detected.get('confidence', 0)
            
            if encoding and confidence > 0.7:
                try:
                    return csv_content.decode(encoding, errors='replace')
                except (UnicodeDecodeError, LookupError):
                    pass
        except Exception:
            pass
    
    # Try common encodings
    encodings_to_try = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
    for encoding in encodings_to_try:
        try:
            decoded = csv_content.decode(encoding, errors='strict')
            if len(decoded) > 0:
                printable_ratio = sum(1 for c in decoded[:1000] if c.isprintable() or c.isspace()) / min(len(decoded), 1000)
                if printable_ratio > 0.7:
                    return decoded
        except (UnicodeDecodeError, LookupError):
            continue
    
    # Fallback to UTF-8 with error replacement
    decoded = csv_content.decode('utf-8', errors='replace')
    printable_ratio = sum(1 for c in decoded[:1000] if c.isprintable() or c.isspace()) / min(len(decoded), 1000)
    if printable_ratio < 0.5:
        raise ValueError(
            f"CSV file {file_name} appears to be corrupted or in an unsupported encoding. "
            f"Please verify the file is a valid CSV file."
        )
    return decoded


def extract_tsv_text(tsv_content: bytes, file_name: str) -> str:
    """Extract text from TSV (Tab-Separated Values) file."""
    # TSV is similar to CSV but uses tabs
    return extract_csv_text(tsv_content, file_name)


def process_image_file(content: Union[str, bytes]) -> str:
    """
    Process image file and return base64 encoded content for Vision GPT.
    
    Args:
        content: Image content as bytes (from FormData) or base64 string (backward compatibility)
    
    Returns:
        Base64 encoded image content (ready for Vision GPT)
    """
    try:
        if isinstance(content, bytes):
            image_data = content
            base64_content = base64.b64encode(image_data).decode('utf-8')
        else:
            if content.startswith('data:'):
                content = content.split(',', 1)[1] if ',' in content else content
            image_data = base64.b64decode(content)
            base64_content = content
        
        # Validate it's a valid image using PIL
        try:
            image = Image.open(io.BytesIO(image_data))
            image.verify()
        except Exception:
            pass  # Continue anyway, let Vision GPT handle it
        
        return base64_content
    except Exception as e:
        raise ValueError(f"Failed to process image: {str(e)}")


def process_attachment(
    attachment_id: str,
    file_name: str,
    file_type: str,
    file_size: int,
    file_type_classification: str,
    content: Optional[Union[str, bytes]]
) -> Tuple[str, str]:
    """
    Process a file attachment based on its type.
    
    Args:
        attachment_id: Unique identifier for the attachment
        file_name: Name of the file
        file_type: MIME type
        file_size: Size in bytes
        file_type_classification: 'image' or 'text'
        content: File content as bytes (from FormData) or base64 string (for backward compatibility)
    
    Returns:
        Tuple of (processed_content, content_type) where:
        - processed_content: Extracted text (for text files) or base64 (for images)
        - content_type: 'text' or 'image'
    """
    if not content:
        raise ValueError(f"File {file_name} has no content")
    
    if file_type_classification == "image":
        processed_content = process_image_file(content)
        return processed_content, "image"
    else:
        # Handle PDF separately for backward compatibility with base64 strings
        if file_type == "application/pdf" or file_name.lower().endswith('.pdf'):
            if isinstance(content, bytes):
                decoded_content = content
            else:
                decoded_content = _decode_content(content, file_name)
            
            processed_content = extract_pdf_text(decoded_content)
            if not processed_content or not processed_content.strip():
                raise ValueError(f"No text content found in PDF {file_name}. The PDF may be image-based or corrupted.")
            return processed_content, "text"
        else:
            processed_content = process_text_file(content, file_type, file_name)
            if not processed_content or not processed_content.strip():
                raise ValueError(f"No readable content found in {file_name}. The file may be empty or corrupted.")
            return processed_content, "text"
